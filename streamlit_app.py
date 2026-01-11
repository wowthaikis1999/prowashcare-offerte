import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table as RLTable, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib import utils
from datetime import datetime
import io
import os

st.set_page_config(page_title="ProWashCare Offerte", layout="centered")
st.title("🧼 ProWashCare – Offerte Tool")

# ================= BEDRIJF =================
BEDRIJFSNAAM = "ProWashCare"
ADRES = "2930 Brasschaat, Antwerpen"
TELEFOON = "+32 470 87 43 39"
EMAIL = "dennisg@prowashcare.com"
WEBSITE = "www.prowashcare.com"

VERVOERSKOSTEN = 8.0
OPRIT_MINIMUM = 299.0

# ================= SESSION =================
if "diensten" not in st.session_state:
    st.session_state.diensten = []

# ================= FUNCTIES =================
def bereken_totaal():
    diensten_clean = [d for d in st.session_state.diensten if not d[0].startswith("Minimumtarief")]

    oprit_diensten = [d for d in diensten_clean if d[0].startswith(("Oprit", "Terras", "Bedrijfsterrein"))]
    oprit_totaal = sum(d[1] for d in oprit_diensten)

    verschil = max(0, OPRIT_MINIMUM - oprit_totaal) if oprit_diensten else 0
    if verschil > 0:
        diensten_clean.append(("Minimumtarief oprit/terras/bedrijfsterrein", verschil))

    subtotaal = sum(d[1] for d in diensten_clean)
    btw = subtotaal * 0.21
    totaal = subtotaal + btw

    return diensten_clean, subtotaal, btw, totaal


def maak_pdf_en_excel(diensten, naam, adres, email, subtotaal, btw, totaal):
    nu = datetime.now()
    nummer = nu.strftime("PWC%Y%m%d%H%M")
    datum = nu.strftime("%d-%m-%Y")

    # ===== EXCEL =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Offerte"

    bold = Font(bold=True)
    ws["A1"] = "ProWashCare – OFFERTE"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:B1")

    ws["A3"] = f"Naam: {naam}"
    ws["A4"] = f"Adres: {adres}"
    ws["A5"] = f"E-mail: {email}"
    ws["A6"] = f"Datum: {datum}"
    ws["A7"] = f"Offertenummer: {nummer}"

    ws["A9"] = "Omschrijving"
    ws["B9"] = "Bedrag"
    ws["A9"].font = bold
    ws["B9"].font = bold

    row = 10
    for oms, bed in diensten:
        ws[f"A{row}"] = oms
        ws[f"B{row}"] = bed
        row += 1

    ws[f"A{row}"] = "Subtotaal"
    ws[f"B{row}"] = subtotaal
    row += 1
    ws[f"A{row}"] = "BTW 21%"
    ws[f"B{row}"] = btw
    row += 1
    ws[f"A{row}"] = "Totaal"
    ws[f"B{row}"] = totaal

    excel = io.BytesIO()
    wb.save(excel)
    excel.seek(0)

    # ===== PDF =====
    pdf_buf = io.BytesIO()
    pdf = SimpleDocTemplate(pdf_buf)
    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("<b>ProWashCare – Offerte</b>", styles["Title"]))
    content.append(Spacer(1, 12))
    content.append(Paragraph(f"<b>Naam:</b> {naam}", styles["Normal"]))
    content.append(Paragraph(f"<b>Adres:</b> {adres}", styles["Normal"]))
    content.append(Paragraph(f"<b>E-mail:</b> {email}", styles["Normal"]))
    content.append(Paragraph(f"<b>Datum:</b> {datum}", styles["Normal"]))
    content.append(Paragraph(f"<b>Offertenummer:</b> {nummer}", styles["Normal"]))
    content.append(Spacer(1, 20))

    table_data = [["Omschrijving", "Bedrag (€)"]]
    for d in diensten:
        table_data.append([Paragraph(d[0].replace("\n", "<br/>"), styles["Normal"]), f"{d[1]:.2f}"])

    table_data += [
        ["", ""],
        ["Subtotaal", f"{subtotaal:.2f}"],
        ["BTW 21%", f"{btw:.2f}"],
        ["Totaal", f"{totaal:.2f}"],
    ]

    content.append(RLTable(table_data, colWidths=[350, 100]))
    pdf.build(content)
    pdf_buf.seek(0)

    return excel, pdf_buf, nummer

# ================= KLANT =================
st.write("### Klantgegevens")
c1, c2 = st.columns(2)
with c1:
    klant_naam = st.text_input("Naam")
    klant_email = st.text_input("E-mail")
with c2:
    klant_adres = st.text_area("Adres", height=80)

# ================= DIENSTEN =================
st.write("### Dienst kiezen")
dienst = st.selectbox("Dienst", [
    "Ramen wassen", "Zonnepanelen", "Gevelreiniging", "Oprit / Terras / Bedrijfsterrein"
], label_visibility="collapsed")

omschrijving = dienst
bedrag = 0

if dienst == "Ramen wassen":
    kb = st.number_input("Kleine ramen – Binnen", 0)
    kbu = st.number_input("Kleine ramen – Buiten", 0)
    gb = st.number_input("Grote ramen – Binnen", 0)
    gbu = st.number_input("Grote ramen – Buiten", 0)
    dakb = st.number_input("Dakramen / moeilijk te bereiken raam – Binnen", 0)
    dakbu = st.number_input("Dakramen / moeilijk te bereiken raam – Buiten", 0)

    regels = []
    if kb: regels.append(f"- Kleine ramen binnen: {kb} × €2.00 = €{kb*2:.2f}")
    if kbu: regels.append(f"- Kleine ramen buiten: {kbu} × €1.50 = €{kbu*1.5:.2f}")
    if gb: regels.append(f"- Grote ramen binnen: {gb} × €2.50 = €{gb*2.5:.2f}")
    if gbu: regels.append(f"- Grote ramen buiten: {gbu} × €2.00 = €{gbu*2:.2f}")
    if dakb: regels.append(f"- Dakramen binnen: {dakb} × €2.50 = €{dakb*2.5:.2f}")
    if dakbu: regels.append(f"- Dakramen buiten: {dakbu} × €2.50 = €{dakbu*2.5:.2f}")

    bedrag = kb*2 + kbu*1.5 + gb*2.5 + gbu*2 + dakb*2.5 + dakbu*2.5
    omschrijving += "\n" + "\n".join(regels)

elif dienst == "Zonnepanelen":
    aantal = st.number_input("Aantal panelen", 1)
    bedrag = max(79, aantal * 5)
    omschrijving += f"\n- {aantal} panelen × €5.00"

elif dienst == "Gevelreiniging":
    m2 = st.number_input("m²", 0.1)
    bedrag = max(299, m2 * 5)
    omschrijving += f"\n- {m2} m² × €5.00"

elif dienst == "Oprit / Terras / Bedrijfsterrein":
    type_keuze = st.radio("Type", ["Oprit", "Terras", "Bedrijfsterrein"], horizontal=True)
    m2 = st.number_input("Oppervlakte (m²)", min_value=0.1, format="%.1f")

    col1, col2, col3, col4 = st.columns(4)
    reinigen = col1.checkbox("Reinigen")
    zand = col2.checkbox("Zand invegen")
    onkruid = col3.checkbox("Onkruidmijdend voegzand")
    coating = col4.checkbox("Coating")

    bedrag = 0
    regels = []

    if reinigen:
        bedrag += m2 * 3.5
        regels.append(f"- Reinigen: {m2} m² × €3,50 = €{m2*3.5:.2f}")
    if zand:
        bedrag += m2 * 1.0
        regels.append(f"- Zand invegen: {m2} m² × €1,00 = €{m2*1.0:.2f}")
    if onkruid:
        bedrag += m2 * 2.0
        regels.append(f"- Onkruidmijdend voegzand: {m2} m² × €2,00 = €{m2*2.0:.2f}")
    if coating:
        bedrag += m2 * 3.5
        regels.append(f"- Coating: {m2} m² × €3,50 = €{m2*3.5:.2f}")

    if regels:
        omschrijving = f"{type_keuze}\n" + "\n".join(regels)
    else:
        st.warning("Selecteer minstens één optie.")
        bedrag = 0


# ================= KNOPPEN =================
col1, col2 = st.columns(2)

with col1:
    if st.button("➕ Dienst toevoegen") and bedrag > 0:
        st.session_state.diensten.append((omschrijving, bedrag))
        st.rerun()

with col2:
    if st.button("🚚 Vervoerskosten toevoegen"):
        if not any(d[0] == "Vervoerskosten" for d in st.session_state.diensten):
            st.session_state.diensten.append(("Vervoerskosten", VERVOERSKOSTEN))
            st.rerun()

# ================= OVERZICHT =================
st.write("### Overzicht")
diensten_final, subtotaal, btw, totaal = bereken_totaal()

for i, d in enumerate(diensten_final):
    c1, c2, c3 = st.columns([6, 2, 1])
    c1.write(d[0].replace("\n", "  \n"))
    c2.write(f"€ {d[1]:.2f}")
    if c3.button("❌", key=i):
        st.session_state.diensten.pop(i)
        st.rerun()

st.write(f"**Totaal incl. btw:** € {totaal:.2f}")

# ================= OFFERTES =================
if st.button("Maak offerte (PDF + Excel)"):
    excel, pdf, nr = maak_pdf_en_excel(
        diensten_final, klant_naam, klant_adres, klant_email, subtotaal, btw, totaal
    )
    st.download_button("📊 Excel", excel, f"Offerte_{nr}.xlsx")
    st.download_button("📄 PDF", pdf, f"Offerte_{nr}.pdf")
